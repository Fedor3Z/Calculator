from __future__ import annotations

from typing import Dict, List

from openpyxl import Workbook

from core.model import InputCell
from solver.optimizer import ConstraintStatus


def export_report(
    path: str,
    inputs: List[InputCell],
    values: Dict[str, float],
    key_outputs: Dict[str, float],
    formulas: Dict[str, str],
    constraints: List[ConstraintStatus],
) -> None:
    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs.append(["Name", "Cell", "Value", "Unit", "Description"])
    for item in inputs:
        ws_inputs.append([item.name, item.cell, values.get(item.cell), item.unit, item.description])

    ws_outputs = wb.create_sheet("Key Outputs")
    ws_outputs.append(["Name", "Cell", "Value", "Unit", "Notes"])
    for cell, value in key_outputs.items():
        ws_outputs.append([cell, cell, value, "", ""])

    ws_formulas = wb.create_sheet("Intermediates")
    ws_formulas.append(["Cell", "Value", "Formula"])
    for cell, formula in formulas.items():
        ws_formulas.append([cell, values.get(cell), formula])

    ws_constraints = wb.create_sheet("Constraints")
    ws_constraints.append(["Constraint", "Violation", "Status"])
    for item in constraints:
        ws_constraints.append([item.name, item.violation, item.status])

    wb.save(path)
